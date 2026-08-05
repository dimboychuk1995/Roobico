"""Импорт/экспорт данных магазина: Customers, Units, Vendors, Parts, Work Orders.

Импорт: CSV/XLSX с ручным маппингом колонок (двухшаговый flow: заголовки →
маппинг → импорт). Каждая пропущенная строка попадает в отчёт с причиной.
Дубли отклоняются теми же правилами, что и создание руками
(app/utils/duplicates.py). Остатки партов проводятся через сервис склада
(apply_stock_change), юниты привязываются к клиенту по имени, work orders
считают тоталы тем же конвейером, что мобильный редактор
(compute_labors_and_totals → align_totals_with_labors).

Экспорт: CSV (utf-8 BOM, дружелюбен к Excel) или XLSX. Колонки экспорта
совпадают с label'ами полей импорта — выгруженный файл маппится обратно
автоматически.

Права: страница — import_export.view; импорт — import_export.import;
экспорт — import_export.export (у owner есть всё).
"""
from __future__ import annotations

import csv
import io
import json
from datetime import date as date_cls, datetime, timezone

from bson import ObjectId
from flask import jsonify, request, send_file, session

from app.blueprints.import_export import import_export_bp
from app.blueprints.main.routes import _render_app_page
from app.extensions import get_master_db, get_mongo_client
from app.utils.auth import (
    SESSION_SHOP_ID,
    SESSION_TENANT_ID,
    SESSION_USER_ID,
    login_required,
)
from app.utils.duplicates import (
    _norm as _dup_norm,
    customer_display_name,
    find_duplicate_customer,
    find_duplicate_part,
    find_duplicate_unit,
    find_duplicate_vendor,
)
from app.utils.entity_search import build_customer_search_terms, build_unit_search_terms
from app.utils.mongo_tx import run_atomically
from app.utils.parts_search import build_parts_search_terms
from app.utils.permissions import permission_required


def _oid(value):
    if not value:
        return None
    try:
        return ObjectId(str(value))
    except Exception:
        return None


def utcnow():
    return datetime.now(timezone.utc)


# ── field definitions per entity ─────────────────────────────────────

ENTITY_FIELDS = {
    "customers": [
        {"key": "company_name", "label": "Company Name"},
        {"key": "first_name", "label": "First Name"},
        {"key": "last_name", "label": "Last Name"},
        {"key": "phone", "label": "Phone"},
        {"key": "email", "label": "Email"},
        {"key": "address", "label": "Address"},
        {"key": "pricing_rule_name", "label": "Pricing Scale Name"},
    ],
    "units": [
        {"key": "customer_name", "label": "Customer Name"},
        {"key": "unit_number", "label": "Unit Number"},
        {"key": "vin", "label": "VIN"},
        {"key": "year", "label": "Year"},
        {"key": "make", "label": "Make"},
        {"key": "model", "label": "Model"},
        {"key": "type", "label": "Type"},
        {"key": "mileage", "label": "Mileage"},
    ],
    "vendors": [
        {"key": "name", "label": "Vendor Name"},
        {"key": "first_name", "label": "Contact First Name"},
        {"key": "last_name", "label": "Contact Last Name"},
        {"key": "phone", "label": "Phone"},
        {"key": "email", "label": "Email"},
        {"key": "website", "label": "Website"},
        {"key": "address", "label": "Address"},
        {"key": "notes", "label": "Notes"},
    ],
    "parts": [
        {"key": "part_number", "label": "Part Number"},
        {"key": "description", "label": "Description"},
        {"key": "reference", "label": "Reference"},
        {"key": "in_stock", "label": "In Stock"},
        {"key": "average_cost", "label": "Average Cost"},
        {"key": "selling_price", "label": "Selling Price"},
    ],
    "work_orders": [
        {"key": "wo_number", "label": "WO Number"},
        {"key": "date", "label": "Date"},
        {"key": "customer_name", "label": "Customer Name"},
        {"key": "unit_number", "label": "Unit Number"},
        {"key": "vin", "label": "VIN"},
        {"key": "mileage", "label": "Mileage"},
        {"key": "status", "label": "Status"},
        {"key": "description", "label": "Description"},
        {"key": "hours", "label": "Hours"},
        {"key": "labor_total", "label": "Labor Total"},
        {"key": "parts_total", "label": "Parts Total"},
        {"key": "sales_tax_total", "label": "Sales Tax"},
        {"key": "paid_amount", "label": "Paid Amount"},
    ],
}

ENTITY_LABELS = {
    "customers": "Customers",
    "units": "Units",
    "vendors": "Vendors",
    "parts": "Parts",
    "work_orders": "Work Orders",
}

WO_STATUSES = {"open", "in_progress", "completed", "paid"}
MAX_ERRORS = 25


# ── helpers ──────────────────────────────────────────────────────────


def _get_shop_db():
    master = get_master_db()
    tenant_id = _oid(session.get(SESSION_TENANT_ID))
    shop_id = _oid(session.get(SESSION_SHOP_ID))
    if not tenant_id or not shop_id:
        return None, None

    shop = master.shops.find_one({"_id": shop_id, "tenant_id": tenant_id, "is_active": True})
    if not shop:
        return None, None

    db_name = shop.get("db_name")
    if not db_name:
        return None, None

    client = get_mongo_client()
    return client[db_name], shop


def _check_import_file(file_storage):
    """Единая валидация загруженного файла. Возвращает текст ошибки или None."""
    if not file_storage or not file_storage.filename:
        return "No file uploaded."
    fname = file_storage.filename.lower()
    if fname.endswith(".xls") and not fname.endswith(".xlsx"):
        return "Legacy .xls files are not supported. Save the file as .xlsx or CSV and try again."
    if not fname.endswith((".csv", ".xlsx")):
        return "Unsupported file format. Use CSV or Excel (.xlsx)."
    return None


def _parse_all_rows(file_storage):
    """(headers, rows) — заголовки страйпятся ОДИНАКОВО для обоих форматов
    (расхождение strip'а между шагами и было причиной «0 imported» на CSV
    с пробелами после запятых)."""
    filename = (file_storage.filename or "").lower()

    if filename.endswith(".csv"):
        raw = file_storage.read()
        file_storage.seek(0)
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
    else:  # .xlsx
        import openpyxl

        wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        file_storage.seek(0)

    if not all_rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
    rows = []
    for row_vals in all_rows[1:]:
        row_dict = {}
        for i, h in enumerate(headers):
            if h:
                row_dict[h] = row_vals[i] if i < len(row_vals) else None
        if any(v is not None and str(v).strip() for v in row_dict.values()):
            rows.append(row_dict)
    return [h for h in headers if h], rows


def _clean_excel(val):
    """Strip Excel ='"..."' and ="..." wrappers from a value."""
    if val is None:
        return None
    s = str(val).strip()
    if (s.startswith('="') and s.endswith('"')) or (s.startswith("='") and s.endswith("'")):
        s = s[2:-1]
    return s


def _safe_str(val):
    if val is None:
        return None
    s = _clean_excel(val)
    s = s.strip() if s else None
    return s if s else None


def _safe_int(val):
    if val is None or str(val).strip() == "":
        return None
    try:
        s = _clean_excel(val).replace("$", "").replace(",", "").strip()
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _safe_float(val):
    if val is None or str(val).strip() == "":
        return None
    try:
        s = _clean_excel(val).replace("$", "").replace(",", "").strip()
        return round(float(s), 2)
    except (ValueError, TypeError):
        return None


def _safe_date(val):
    """datetime из ячейки: datetime/date из xlsx либо строка в ходовых форматах."""
    if isinstance(val, datetime):
        return val if val.tzinfo else val.replace(tzinfo=timezone.utc)
    if isinstance(val, date_cls):
        return datetime(val.year, val.month, val.day, tzinfo=timezone.utc)
    s = _safe_str(val)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _active_customer_map(shop_db, shop_id):
    """_norm(отображаемое имя) -> customer _id, только активные."""
    out = {}
    for c in shop_db.customers.find(
        {"shop_id": shop_id, "is_active": {"$ne": False}},
        {"company_name": 1, "contacts": 1, "first_name": 1, "last_name": 1},
    ):
        key = _dup_norm(customer_display_name(c))
        if key:
            out.setdefault(key, c["_id"])
    return out


# ── doc builders (return (doc | None, reason | None)) ────────────────


def _build_customer_doc(mapped_row, shop, now, user_id, default_labor_rate_id=None,
                        pricing_rule_lookup=None, default_pricing_rule_id=None):
    company_name = _safe_str(mapped_row.get("company_name"))
    first_name = _safe_str(mapped_row.get("first_name"))
    last_name = _safe_str(mapped_row.get("last_name"))
    phone = _safe_str(mapped_row.get("phone"))
    email = _safe_str(mapped_row.get("email"))
    address = _safe_str(mapped_row.get("address"))
    pricing_rule_name = _safe_str(mapped_row.get("pricing_rule_name"))

    if not company_name and not first_name and not last_name:
        return None, "no company name and no contact name"

    contacts = []
    if first_name or last_name or phone or email:
        contacts.append({
            "first_name": first_name or "",
            "last_name": last_name or "",
            "phone": phone or "",
            "email": (email or "").lower(),
            "is_main": True,
        })

    doc = {
        "company_name": company_name,
        "contacts": contacts,
        "address": address,
        "first_name": first_name,
        "last_name": last_name,
        "phone": phone,
        "email": (email or "").lower() if email else None,
        "main_contact_name": " ".join(filter(None, [first_name, last_name])) or None,
        "main_contact_phone": phone,
        "main_contact_email": (email or "").lower() if email else None,
        "taxable": False,
        "current_balance": 0.0,
        "default_labor_rate": default_labor_rate_id,
        "pricing_rule_id": (
            (pricing_rule_lookup or {}).get((pricing_rule_name or "").strip().lower())
            if pricing_rule_name else None
        ) or default_pricing_rule_id,
        "override_part_selling_price": False,
        "is_active": True,
        "shop_id": shop["_id"],
        "tenant_id": shop.get("tenant_id"),
        "created_at": now,
        "updated_at": now,
        "created_by": user_id,
        "updated_by": user_id,
    }
    doc["search_terms"] = build_customer_search_terms(doc)
    return doc, None


def _build_unit_doc(mapped_row, customer_id, shop, now, user_id):
    unit_number = _safe_str(mapped_row.get("unit_number"))
    vin = _safe_str(mapped_row.get("vin"))
    if not unit_number and not vin:
        return None, "unit number and VIN are both empty"

    doc = {
        "customer_id": customer_id,
        "unit_number": unit_number,
        "vin": vin.upper() if vin else None,
        "year": _safe_int(mapped_row.get("year")),
        "make": _safe_str(mapped_row.get("make")),
        "model": _safe_str(mapped_row.get("model")),
        "type": _safe_str(mapped_row.get("type")),
        "mileage": _safe_int(mapped_row.get("mileage")),
        "is_active": True,
        "shop_id": shop["_id"],
        "tenant_id": shop.get("tenant_id"),
        "created_at": now,
        "updated_at": now,
        "created_by": user_id,
        "updated_by": user_id,
    }
    doc["search_terms"] = build_unit_search_terms(doc)
    return doc, None


def _build_vendor_doc(mapped_row, shop, now, user_id):
    name = _safe_str(mapped_row.get("name"))
    if not name:
        return None, "vendor name is empty"

    first_name = _safe_str(mapped_row.get("first_name"))
    last_name = _safe_str(mapped_row.get("last_name"))
    phone = _safe_str(mapped_row.get("phone"))
    email = _safe_str(mapped_row.get("email"))

    contacts = []
    if first_name or last_name or phone or email:
        contacts.append({
            "first_name": first_name or "",
            "last_name": last_name or "",
            "phone": phone or "",
            "email": (email or "").lower(),
            "is_main": True,
        })

    doc = {
        "name": name,
        "website": _safe_str(mapped_row.get("website")),
        "address": _safe_str(mapped_row.get("address")),
        "notes": _safe_str(mapped_row.get("notes")),
        "contacts": contacts,
        "primary_contact_first_name": first_name,
        "primary_contact_last_name": last_name,
        "phone": phone,
        "email": (email or "").lower() if email else None,
        "is_active": True,
        "shop_id": shop["_id"],
        "tenant_id": shop.get("tenant_id"),
        "created_at": now,
        "updated_at": now,
        "created_by": user_id,
        "updated_by": user_id,
    }
    return doc, None


def _build_part_doc(mapped_row, shop, now, user_id):
    part_number = _safe_str(mapped_row.get("part_number"))
    if not part_number:
        return None, "part number is empty"

    description = _safe_str(mapped_row.get("description"))
    reference = _safe_str(mapped_row.get("reference"))
    average_cost = _safe_float(mapped_row.get("average_cost"))
    selling_price = _safe_float(mapped_row.get("selling_price"))

    doc = {
        "part_number": part_number,
        "description": description,
        "reference": reference,
        "search_terms": build_parts_search_terms(part_number, description, reference),
        # Остаток проводится ПОСЛЕ вставки через apply_stock_change (initial),
        # чтобы появились строка локации и движение склада.
        "in_stock": 0,
        "average_cost": average_cost or 0.0,
        "has_selling_price": selling_price is not None and selling_price > 0,
        "selling_price": selling_price,
        "do_not_track_inventory": False,
        "core_has_charge": False,
        "core_cost": None,
        "misc_has_charge": False,
        "misc_charges": [],
        "vendor_id": None,
        "category_id": None,
        "location_id": None,
        "is_active": True,
        "shop_id": shop["_id"],
        "tenant_id": shop.get("tenant_id"),
        "created_at": now,
        "updated_at": now,
        "created_by": user_id,
        "updated_by": user_id,
    }
    return doc, None


def _import_work_order(shop_db, shop, mapped_row, ctx, now, user_id):
    """Импорт одного WO. Возвращает (wo_id | None, reason | None).

    Тоталы считаются тем же конвейером, что и мобильный редактор:
    compute_labors_and_totals → align_totals_with_labors. Parts Total
    представляется one-time-строкой (инвентарь не трогается). Для paid-WO
    платёж на всю сумму пишется атомарно вместе с WO (run_atomically) —
    иначе Outstanding Balance посчитал бы его как долг.
    """
    from app.blueprints.work_orders.services.mobile_editor import compute_labors_and_totals
    from app.blueprints.work_orders.services.totals import (
        align_totals_with_labors,
        get_next_wo_number,
        normalize_totals_payload,
    )

    wo_date = _safe_date(mapped_row.get("date"))
    if not wo_date:
        return None, "date is missing or not recognized (use YYYY-MM-DD or MM/DD/YYYY)"

    customer_name = _safe_str(mapped_row.get("customer_name"))
    if not customer_name:
        return None, "customer name is empty"
    customer_id = ctx["customer_map"].get(_dup_norm(customer_name))
    if not customer_id:
        return None, f'customer "{customer_name}" not found (import customers first)'

    status = (_safe_str(mapped_row.get("status")) or "completed").lower().replace(" ", "_")
    if status not in WO_STATUSES:
        return None, f'status "{status}" is not one of: {", ".join(sorted(WO_STATUSES))}'

    # WO number: свой (с проверкой дублей) либо автоматический
    wo_number = _safe_int(mapped_row.get("wo_number"))
    if wo_number is not None:
        if wo_number in ctx["seen_wo_numbers"] or shop_db.work_orders.count_documents(
            {"shop_id": shop["_id"], "wo_number": wo_number}
        ):
            return None, f"WO #{wo_number} already exists"

    # Юнит (опционально): по номеру или VIN в рамках клиента
    unit_id = None
    unit_number = _safe_str(mapped_row.get("unit_number"))
    vin = _safe_str(mapped_row.get("vin"))
    if unit_number or vin:
        ors = []
        if unit_number:
            ors.append({"unit_number": {"$regex": rf"^\s*{_re_escape(unit_number)}\s*$", "$options": "i"}})
        if vin:
            ors.append({"vin": {"$regex": rf"^\s*{_re_escape(vin)}\s*$", "$options": "i"}})
        unit = shop_db.units.find_one(
            {"shop_id": shop["_id"], "customer_id": customer_id, "$or": ors},
            {"_id": 1},
        )
        if not unit:
            return None, (f'unit "{unit_number or vin}" not found for customer '
                          f'"{customer_name}" (import units first)')
        unit_id = unit["_id"]

    description = _safe_str(mapped_row.get("description")) or "Imported work"
    hours = _safe_float(mapped_row.get("hours"))
    labor_total = _safe_float(mapped_row.get("labor_total"))
    parts_total = _safe_float(mapped_row.get("parts_total"))
    sales_tax_total = _safe_float(mapped_row.get("sales_tax_total")) or 0.0

    parts_payload = []
    if parts_total and parts_total > 0:
        parts_payload.append({
            "one_time_part": True,
            "description": "Imported parts",
            "qty": 1,
            "price": parts_total,
            "cost": 0,
        })

    labors_payload = [{
        "description": description,
        "hours": str(hours) if hours else "",
        "rate_code": ctx["default_rate_code"] if (hours and not labor_total) else "",
        "labor_total": labor_total,
        "parts": parts_payload,
    }]
    labors, totals_raw = compute_labors_and_totals(shop_db, shop, labors_payload)
    # Суммы исторические: shop supply текущего магазина к ним не применяем,
    # иначе labor total уедет от значений из файла.
    from app.blueprints.work_orders.services.common import round2

    for idx, block in enumerate(totals_raw.get("labors") or []):
        supply = round2(block.get("shop_supply_total") or 0)
        if supply:
            block["labor_full_total"] = round2(round2(block.get("labor_full_total") or 0) - supply)
            block["shop_supply_total"] = 0.0
            labor_info = (labors[idx] or {}).get("labor") if idx < len(labors) else None
            if isinstance(labor_info, dict):
                labor_info["labor_full_total"] = block["labor_full_total"]
    totals_raw["shop_supply_total"] = 0.0
    totals_raw["sales_tax_total"] = sales_tax_total
    totals_raw["is_taxable"] = sales_tax_total > 0
    totals = align_totals_with_labors(normalize_totals_payload(totals_raw), labors)

    if wo_number is None:
        wo_number = get_next_wo_number(shop_db, shop["_id"])
    ctx["seen_wo_numbers"].add(wo_number)

    doc = {
        "shop_id": shop["_id"],
        "tenant_id": shop.get("tenant_id"),
        "customer_id": customer_id,
        "unit_id": unit_id,
        "wo_number": wo_number,
        "work_order_date": wo_date,
        "labors": labors,
        "totals": totals,
        "status": status,
        "mileage": _safe_int(mapped_row.get("mileage")),
        "is_active": True,
        "created_at": wo_date,
        "updated_at": now,
        "created_by": user_id,
        "updated_by": user_id,
        "mechanic_done": False,
        "manager_confirmed": False,
        "imported": True,
    }

    if status != "paid":
        return shop_db.work_orders.insert_one(doc).inserted_id, None

    paid_amount = _safe_float(mapped_row.get("paid_amount"))
    amount = paid_amount if paid_amount is not None else (totals.get("grand_total") or 0.0)

    def _tx(tx_session):
        wo_id = shop_db.work_orders.insert_one(doc, session=tx_session).inserted_id
        shop_db.work_order_payments.insert_one({
            "shop_id": shop["_id"],
            "tenant_id": shop.get("tenant_id"),
            "work_order_id": wo_id,
            "amount": float(amount),
            "payment_method": "imported",
            "payment_date": wo_date,
            "notes": "Imported from previous system",
            "is_active": True,
            "created_at": now,
            "created_by": user_id,
        }, session=tx_session)
        return wo_id

    return run_atomically(get_mongo_client(), _tx), None


def _re_escape(value: str) -> str:
    import re

    return re.escape(str(value or "").strip())


# ── routes ───────────────────────────────────────────────────────────


@import_export_bp.get("/")
@login_required
@permission_required("import_export.view")
def import_export_index():
    tab = (request.args.get("tab") or "customers").strip().lower()
    if tab not in ENTITY_LABELS:
        tab = "customers"

    return _render_app_page(
        "public/import_export.html",
        active_page="import_export",
        active_tab=tab,
        entity_tabs=ENTITY_LABELS,
        entity_fields=ENTITY_FIELDS.get(tab, []),
        entity_fields_json=json.dumps(ENTITY_FIELDS.get(tab, [])),
    )


@import_export_bp.post("/upload-headers")
@login_required
@permission_required("import_export.import")
def upload_headers():
    """Parse uploaded file and return headers as JSON."""
    f = request.files.get("file")
    err = _check_import_file(f)
    if err:
        return jsonify({"ok": False, "error": err}), 400

    try:
        headers, _ = _parse_all_rows(f)
    except Exception:
        return jsonify({"ok": False, "error": "Could not read the file. Make sure it is a valid CSV or .xlsx."}), 400
    if not headers:
        return jsonify({"ok": False, "error": "No headers found in the file."}), 400

    return jsonify({"ok": True, "headers": headers})


@import_export_bp.post("/import")
@login_required
@permission_required("import_export.import")
def run_import():
    """Execute the import with field mapping."""
    shop_db, shop = _get_shop_db()
    if shop_db is None:
        return jsonify({"ok": False, "error": "Shop not configured."}), 400

    entity_type = (request.form.get("entity_type") or "").strip()
    if entity_type not in ENTITY_LABELS:
        return jsonify({"ok": False, "error": "Invalid entity type."}), 400

    try:
        mapping = json.loads(request.form.get("mapping") or "{}")
    except (json.JSONDecodeError, TypeError):
        return jsonify({"ok": False, "error": "Invalid field mapping."}), 400
    if not mapping:
        return jsonify({"ok": False, "error": "No fields mapped."}), 400

    f = request.files.get("file")
    file_err = _check_import_file(f)
    if file_err:
        return jsonify({"ok": False, "error": file_err}), 400

    try:
        _, rows = _parse_all_rows(f)
    except Exception:
        return jsonify({"ok": False, "error": "Could not read the file. Make sure it is a valid CSV or .xlsx."}), 400
    if not rows:
        return jsonify({"ok": False, "error": "No data rows found."}), 400

    now = utcnow()
    user_id = _oid(session.get(SESSION_USER_ID))

    imported = 0
    skipped = 0
    errors = []

    def _err(i, reason):
        nonlocal skipped
        skipped += 1
        if len(errors) < MAX_ERRORS:
            errors.append(f"Row {i + 2}: {reason}")

    # ── контекст для конкретной сущности ────────────────────────────
    customer_map = {}
    if entity_type in ("units", "work_orders"):
        customer_map = _active_customer_map(shop_db, shop["_id"])

    customer_default_rate_id = None
    pricing_rule_lookup = {}
    customer_default_pricing_rule_id = None
    if entity_type == "customers":
        from app.blueprints.customers.routes import (
            _resolve_default_labor_rate_id,
            _resolve_default_pricing_rule_id,
        )
        customer_default_rate_id = _resolve_default_labor_rate_id(shop_db, shop["_id"])
        if not customer_default_rate_id:
            return jsonify({"ok": False, "error": "No labor rates configured for this shop. Please create at least one labor rate first."}), 400
        customer_default_pricing_rule_id = _resolve_default_pricing_rule_id(shop_db, shop["_id"])
        for s in shop_db.parts_pricing_rules.find({"shop_id": shop["_id"]}, {"_id": 1, "name": 1}):
            nm = (s.get("name") or "").strip().lower()
            if nm:
                pricing_rule_lookup[nm] = s["_id"]

    wo_ctx = None
    if entity_type == "work_orders":
        from app.blueprints.work_orders.services.lookups import get_labor_rates

        rates = get_labor_rates(shop_db, shop["_id"])
        wo_ctx = {
            "customer_map": customer_map,
            "default_rate_code": (rates[0]["code"] if rates else ""),
            "seen_wo_numbers": set(),
        }

    seen_in_file: set = set()

    for i, row in enumerate(rows):
        mapped_row = {}
        for file_header, our_key in mapping.items():
            if our_key and file_header in row:
                mapped_row[our_key] = row[file_header]

        try:
            if entity_type == "customers":
                doc, reason = _build_customer_doc(
                    mapped_row, shop, now, user_id,
                    default_labor_rate_id=customer_default_rate_id,
                    pricing_rule_lookup=pricing_rule_lookup,
                    default_pricing_rule_id=customer_default_pricing_rule_id,
                )
                if doc is None:
                    _err(i, reason)
                    continue
                label = customer_display_name(doc)
                existing = find_duplicate_customer(
                    shop_db, shop["_id"], doc.get("company_name"), doc.get("contacts"))
                if existing or _dup_norm(label) in seen_in_file:
                    note = " (deactivated)" if existing and existing.get("is_active") is False else ""
                    _err(i, f'customer "{label}" already exists{note}')
                    continue
                seen_in_file.add(_dup_norm(label))
                shop_db.customers.insert_one(doc)
                imported += 1

            elif entity_type == "units":
                customer_name = _safe_str(mapped_row.get("customer_name"))
                customer_id = customer_map.get(_dup_norm(customer_name)) if customer_name else None
                if customer_name and not customer_id:
                    _err(i, f'customer "{customer_name}" not found (import customers first)')
                    continue
                doc, reason = _build_unit_doc(mapped_row, customer_id, shop, now, user_id)
                if doc is None:
                    _err(i, reason)
                    continue
                vin = doc.get("vin")
                file_key = ("unit", str(customer_id), _dup_norm(vin)) if vin else None
                if customer_id and vin:
                    existing = find_duplicate_unit(shop_db, shop["_id"], customer_id, vin)
                    if existing or file_key in seen_in_file:
                        _err(i, f'unit with VIN "{vin}" already exists for customer "{customer_name}"')
                        continue
                if file_key:
                    seen_in_file.add(file_key)
                shop_db.units.insert_one(doc)
                imported += 1

            elif entity_type == "vendors":
                doc, reason = _build_vendor_doc(mapped_row, shop, now, user_id)
                if doc is None:
                    _err(i, reason)
                    continue
                existing = find_duplicate_vendor(shop_db, shop["_id"], doc.get("name"))
                if existing or _dup_norm(doc.get("name")) in seen_in_file:
                    note = " (deactivated)" if existing and existing.get("is_active") is False else ""
                    _err(i, f'vendor "{doc.get("name")}" already exists{note}')
                    continue
                seen_in_file.add(_dup_norm(doc.get("name")))
                shop_db.vendors.insert_one(doc)
                imported += 1

            elif entity_type == "parts":
                doc, reason = _build_part_doc(mapped_row, shop, now, user_id)
                if doc is None:
                    _err(i, reason)
                    continue
                existing = find_duplicate_part(shop_db, shop["_id"], doc.get("part_number"))
                if existing or _dup_norm(doc.get("part_number")) in seen_in_file:
                    note = " (deactivated)" if existing and existing.get("is_active") is False else ""
                    _err(i, f'part "{doc.get("part_number")}" already exists{note}')
                    continue
                seen_in_file.add(_dup_norm(doc.get("part_number")))
                in_stock = _safe_int(mapped_row.get("in_stock"))
                res = shop_db.parts.insert_one(doc)
                if in_stock:
                    # Стартовый остаток — через сервис склада, как ручное
                    # создание: строка локации + движение "initial".
                    from app.blueprints.parts.services.stock import apply_stock_change

                    apply_stock_change(
                        shop_db, shop["_id"], res.inserted_id, in_stock, "initial",
                        user_id=user_id,
                    )
                imported += 1

            elif entity_type == "work_orders":
                wo_id, reason = _import_work_order(shop_db, shop, mapped_row, wo_ctx, now, user_id)
                if wo_id is None:
                    _err(i, reason)
                    continue
                imported += 1

        except Exception as exc:  # noqa: BLE001
            _err(i, str(exc))

    result = {
        "ok": True,
        "imported": imported,
        "skipped": skipped,
        "total": len(rows),
    }
    if errors:
        result["errors"] = errors
    return jsonify(result)


# ── export ───────────────────────────────────────────────────────────


def _yes_no(value) -> str:
    return "no" if value is False else "yes"


def _main_contact(doc) -> dict:
    contacts = [c for c in doc.get("contacts") or [] if isinstance(c, dict)]
    for pool in ([c for c in contacts if c.get("is_main")], contacts):
        for c in pool:
            return c
    return {}


def _export_customers(shop_db, shop_id):
    rule_names = {r["_id"]: r.get("name") or "" for r in
                  shop_db.parts_pricing_rules.find({"shop_id": shop_id}, {"name": 1})}
    headers = ["Company Name", "First Name", "Last Name", "Phone", "Email",
               "Address", "Pricing Scale Name", "Taxable", "Active"]
    rows = []
    for c in shop_db.customers.find({"shop_id": shop_id}).sort("company_name", 1):
        main = _main_contact(c)
        rows.append([
            c.get("company_name") or "",
            main.get("first_name") or c.get("first_name") or "",
            main.get("last_name") or c.get("last_name") or "",
            main.get("phone") or c.get("phone") or "",
            main.get("email") or c.get("email") or "",
            c.get("address") or "",
            rule_names.get(c.get("pricing_rule_id")) or "",
            _yes_no(c.get("taxable", False) or False),
            _yes_no(c.get("is_active")),
        ])
    return headers, rows


def _export_units(shop_db, shop_id):
    customer_names = {}
    for c in shop_db.customers.find(
        {"shop_id": shop_id},
        {"company_name": 1, "contacts": 1, "first_name": 1, "last_name": 1},
    ):
        customer_names[c["_id"]] = customer_display_name(c)
    headers = ["Customer Name", "Unit Number", "VIN", "Year", "Make", "Model",
               "Type", "Mileage", "Active"]
    rows = []
    for u in shop_db.units.find({"shop_id": shop_id}).sort("unit_number", 1):
        rows.append([
            customer_names.get(u.get("customer_id")) or "",
            u.get("unit_number") or "",
            u.get("vin") or "",
            u.get("year") if u.get("year") is not None else "",
            u.get("make") or "",
            u.get("model") or "",
            u.get("type") or "",
            u.get("mileage") if u.get("mileage") is not None else "",
            _yes_no(u.get("is_active")),
        ])
    return headers, rows


def _export_vendors(shop_db, shop_id):
    headers = ["Vendor Name", "Contact First Name", "Contact Last Name", "Phone",
               "Email", "Website", "Address", "Notes", "Active"]
    rows = []
    for v in shop_db.vendors.find({"shop_id": shop_id}).sort("name", 1):
        main = _main_contact(v)
        rows.append([
            v.get("name") or "",
            main.get("first_name") or v.get("primary_contact_first_name") or "",
            main.get("last_name") or v.get("primary_contact_last_name") or "",
            main.get("phone") or v.get("phone") or "",
            main.get("email") or v.get("email") or "",
            v.get("website") or "",
            v.get("address") or "",
            v.get("notes") or "",
            _yes_no(v.get("is_active")),
        ])
    return headers, rows


def _export_parts(shop_db, shop_id):
    headers = ["Part Number", "Description", "Reference", "In Stock",
               "Average Cost", "Selling Price", "Active"]
    rows = []
    query = {"shop_id": shop_id, "merged_into": {"$exists": False}}
    for p in shop_db.parts.find(query).sort("part_number", 1):
        rows.append([
            p.get("part_number") or "",
            p.get("description") or "",
            p.get("reference") or "",
            p.get("in_stock") if p.get("in_stock") is not None else "",
            p.get("average_cost") if p.get("average_cost") is not None else "",
            p.get("selling_price") if p.get("selling_price") is not None else "",
            _yes_no(p.get("is_active")),
        ])
    return headers, rows


def _export_work_orders(shop_db, shop_id):
    customer_names = {}
    for c in shop_db.customers.find(
        {"shop_id": shop_id},
        {"company_name": 1, "contacts": 1, "first_name": 1, "last_name": 1},
    ):
        customer_names[c["_id"]] = customer_display_name(c)
    units = {u["_id"]: u for u in shop_db.units.find(
        {"shop_id": shop_id}, {"unit_number": 1, "vin": 1})}

    paid_map = {}
    for row in shop_db.work_order_payments.aggregate([
        {"$match": {"shop_id": shop_id, "is_active": True}},
        {"$group": {"_id": "$work_order_id", "paid": {"$sum": "$amount"}}},
    ]):
        paid_map[row["_id"]] = round(float(row.get("paid") or 0), 2)

    headers = ["WO Number", "Date", "Status", "Customer Name", "Unit Number",
               "VIN", "Mileage", "Description", "Labor Total", "Parts Total",
               "Sales Tax", "Grand Total", "Paid Amount", "Balance"]
    rows = []
    for wo in shop_db.work_orders.find(
        {"shop_id": shop_id, "is_active": {"$ne": False}}
    ).sort("wo_number", 1):
        totals = wo.get("totals") if isinstance(wo.get("totals"), dict) else {}
        unit = units.get(wo.get("unit_id")) or {}
        wo_date = wo.get("work_order_date") or wo.get("created_at")
        descriptions = []
        for block in wo.get("labors") or []:
            labor = (block or {}).get("labor") if isinstance(block, dict) else None
            desc = str(((labor or {}).get("description")) or "").strip()
            if desc:
                descriptions.append(desc)
        grand = round(float(totals.get("grand_total") or wo.get("grand_total") or 0), 2)
        paid = paid_map.get(wo.get("_id"), 0.0)
        rows.append([
            wo.get("wo_number") or "",
            wo_date.strftime("%Y-%m-%d") if isinstance(wo_date, datetime) else "",
            wo.get("status") or "",
            customer_names.get(wo.get("customer_id")) or "",
            unit.get("unit_number") or "",
            unit.get("vin") or "",
            wo.get("mileage") if wo.get("mileage") is not None else "",
            "; ".join(descriptions),
            round(float(totals.get("labor_total") or 0), 2),
            round(float(totals.get("parts_total") or 0), 2),
            round(float(totals.get("sales_tax_total") or 0), 2),
            grand,
            paid,
            round(max(0.0, grand - paid), 2),
        ])
    return headers, rows


_EXPORTERS = {
    "customers": _export_customers,
    "units": _export_units,
    "vendors": _export_vendors,
    "parts": _export_parts,
    "work_orders": _export_work_orders,
}


@import_export_bp.get("/export/<entity>")
@login_required
@permission_required("import_export.export")
def run_export(entity):
    """Скачать все записи сущности как CSV (utf-8 BOM) или XLSX."""
    entity = (entity or "").strip().lower()
    exporter = _EXPORTERS.get(entity)
    if exporter is None:
        return jsonify({"ok": False, "error": "Invalid entity type."}), 400

    shop_db, shop = _get_shop_db()
    if shop_db is None:
        return jsonify({"ok": False, "error": "Shop not configured."}), 400

    fmt = (request.args.get("fmt") or "csv").strip().lower()
    if fmt not in ("csv", "xlsx"):
        return jsonify({"ok": False, "error": "Format must be csv or xlsx."}), 400

    headers, rows = exporter(shop_db, shop["_id"])
    stamp = utcnow().strftime("%Y%m%d")
    filename = f"{entity}_{stamp}.{fmt}"

    if fmt == "csv":
        text = io.StringIO()
        writer = csv.writer(text, lineterminator="\r\n")
        writer.writerow(headers)
        writer.writerows(rows)
        # BOM — чтобы Excel открывал utf-8 без кракозябр
        payload = io.BytesIO(("﻿" + text.getvalue()).encode("utf-8"))
        return send_file(payload, mimetype="text/csv; charset=utf-8",
                         as_attachment=True, download_name=filename)

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ENTITY_LABELS.get(entity, entity)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for col_idx, header in enumerate(headers, start=1):
        width = max(len(str(header)) + 2, 12)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    payload = io.BytesIO()
    wb.save(payload)
    payload.seek(0)
    return send_file(
        payload,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=filename)
